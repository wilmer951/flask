import os
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from flask import Flask, request, render_template, flash, redirect, url_for, send_file, session


from flask import Flask, request, render_template, flash, redirect, url_for, send_from_directory, session

app = Flask(__name__)
app.jinja_env.cache = {}

# Configuración para la carpeta de carga
app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
app.config['ALLOWED_EXTENSIONS'] = {'xls', 'xlsx'}
app.secret_key = 'your_secret_key'  # Para mensajes flash


# Ruta para la página de inicio
@app.route('/')
def home():

    return render_template("home.html")


# Función para verificar si el archivo tiene una extensión permitida
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']


# Ruta para la página integracion
@app.route('/integration', methods=['GET', 'POST'])
def integration():
    return render_template("integration.html")


@app.route('/analyticfile', methods=['GET', 'POST'])
def analyticfile():

    status = 0
    if request.method == 'POST':
        file = request.files['file']

        if not (file and allowed_file(file.filename)):

            flash('Solo se permite archivos con extensión .xls o .xlsx')
            status = 0
            return render_template("integration.html", stv=status)

        # Si el archivo es válido
        if file and allowed_file(file.filename):

            # Cargar el archivo Excel directamente en memoria
            file_stream = BytesIO(file.read())
            wb = load_workbook(file_stream)
            ws = wb.active

            # Comprobar que las columnas necesarias están presentes
            required_columns = ['SEXO', 'EDAD', 'REGIMEN', 'VICTIMA DE CONFLICTO',
                                'MIGRANTE', 'CARCELARIO', 'EN ESTADO DE GESTACION',
                                'OTRO', 'DESPLAZADO']

            columns_in_file = [cell.value for cell in ws[1]]
            missing_columns = [
                col for col in required_columns if col not in columns_in_file]

            if missing_columns:

                status = 0
                flash(
                    f'Faltan las siguientes columnas necesarias en el archivo: {", ".join(missing_columns)}', 'error')
                return render_template("integration.html", stv=status)

            # Obtener las filas visibles
            visible_rows = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                if not ws.row_dimensions[row[0].row].hidden:
                    visible_rows.append([cell.value for cell in row])

            # Crear un DataFrame con las filas visibles
            columns = [cell.value for cell in ws[1]]
            df_visible = pd.DataFrame(visible_rows, columns=columns)

            # Modificaciones al DataFrame
            df_visible['HOMBRE'] = ''
            df_visible['MUJER'] = ''
            df_visible['CONTRIBUTIVO'] = ''
            df_visible['SUBSIDIADO'] = ''
            df_visible['CATEGORIA'] = ''
            df_visible['SINCATEGORIA'] = ''

            # Asegurarse de que no haya espacios en blanco ni diferencias de mayúsculas/minúsculas
            df_visible['VICTIMA DE CONFLICTO'] = df_visible['VICTIMA DE CONFLICTO'].str.strip(
            ).str.upper()
            df_visible['MIGRANTE'] = df_visible['MIGRANTE'].str.strip().str.upper()
            df_visible['CARCELARIO'] = df_visible['CARCELARIO'].str.strip(
            ).str.upper()
            df_visible['EN ESTADO DE GESTACION'] = df_visible['EN ESTADO DE GESTACION'].str.strip(
            ).str.upper()
            df_visible['OTRO'] = df_visible['OTRO'].str.strip().str.upper()
            df_visible['DESPLAZADO'] = df_visible['DESPLAZADO'].str.strip(
            ).str.upper()

            # Asignar valores a nuevas columnas
            df_visible.loc[df_visible['SEXO'] == 'M',
                           'HOMBRE'] = df_visible['EDAD'].astype(int)
            df_visible.loc[df_visible['SEXO'] == 'F',
                           'MUJER'] = df_visible['EDAD'].astype(int)
            df_visible.loc[df_visible['REGIMEN'] ==
                           'Contributivo', 'CONTRIBUTIVO'] = 'X'
            df_visible.loc[df_visible['REGIMEN'] ==
                           'Subsidiado', 'SUBSIDIADO'] = 'X'

            for index, row in df_visible.iterrows():
                if (row['VICTIMA DE CONFLICTO'] == 'SI' or
                    row['MIGRANTE'] == 'SI' or
                    row['CARCELARIO'] == 'SI' or
                    row['EN ESTADO DE GESTACION'] == 'SI' or
                        row['OTRO'] == 'SI' or row['DESPLAZADO'] == 'SI'):
                    df_visible.at[index, 'CATEGORIA'] = 'X'
                else:
                    df_visible.at[index, 'SINCATEGORIA'] = 'X'

            # Guardar el DataFrame modificado en un archivo Excel en memoria
            output = BytesIO()
            df_visible.to_excel(output, index=False)
            output.seek(0)

            # Guardar el archivo en la sesión
            # Guardamos el archivo en la sesión
            session['file'] = output.getvalue()

            flash(f'Archivo "{file.filename}" modificdo exitosamente!')
            status = 1

            # Redirigir a la página de descarga

            return render_template("integration.html", stv=status, file=output.getvalue())

    return render_template("integration.html", stv=status)


@app.route('/download_file')
def download_file():
    # Recuperar el archivo modificado de la sesión
    output_data = session.get('file')

    if output_data:
        # Crear un archivo en memoria y devolverlo como archivo descargable
        output = BytesIO(output_data)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name='archivo_modificado.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        flash('No se pudo recuperar el archivo modificado.', 'error')
        return redirect(url_for('home'))


if __name__ == "__main__":
    app.run(debug=True)
