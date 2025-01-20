from flask import flash
from openpyxl import load_workbook
from app.utils import file_origen,file_end
import os
import pandas as pd
from app import app





def analytic_data(request):
    status = 0
    data_file = os.path.join(app.config['UPLOAD_FOLDER'], file_origen)
    wb = load_workbook(data_file)
    ws = wb.active
    required_columns = ['SEXO', 'EDAD', 'REGIMEN', 'VICTIMA DE CONFLICTO', 'MIGRANTE', 'CARCELARIO', 'EN ESTADO DE GESTACION', 'OTRO', 'DESPLAZADO']

    columns_in_file = [cell.value for cell in ws[1]]
    missing_columns = [col for col in required_columns if col not in columns_in_file]

    if missing_columns:
        flash(f'Faltan las siguientes columnas necesarias: {", ".join(missing_columns)}', 'error')
        return status
    
    else:

        visible_rows = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            if not ws.row_dimensions[row[0].row].hidden:
                visible_rows.append([cell.value for cell in row])

        df_visible = pd.DataFrame(visible_rows, columns=[cell.value for cell in ws[1]])

        # Procesamiento de datos (como en tu código original)
      # Asignar las columnas 'HOMBRE' y 'MUJER'
        df_visible['HOMBRE'] = ''
        df_visible['MUJER'] = ''

        df_visible['CONTRIBUTIVO'] = ''
        df_visible['SUBSIDIADO'] = ''

        df_visible['CATEGORIA'] = ''
        df_visible['SINCATEGORIA'] = ''

        # Asegurarse de que no haya espacios en blanco ni diferencias de mayúsculas/minúsculas
        df_visible['VICTIMA DE CONFLICTO'] = df_visible['VICTIMA DE CONFLICTO'].str.strip(
        ).str.upper()
        df_visible['MIGRANTE'] = df_visible['MIGRANTE'].str.strip().str.upper()
        df_visible['CARCELARIO'] = df_visible['CARCELARIO'].str.strip().str.upper()
        df_visible['EN ESTADO DE GESTACION'] = df_visible['EN ESTADO DE GESTACION'].str.strip(
        ).str.upper()
        df_visible['OTRO'] = df_visible['OTRO'].str.strip().str.upper()
        df_visible['DESPLAZADO'] = df_visible['DESPLAZADO'].str.strip().str.upper()

        # Asignar la edad a 'hombres' si el sexo es 'M' (masculino), o a 'mujeres' si el sexo es 'F' (femenino)
        df_visible.loc[df_visible['SEXO'] == 'M',
                       'HOMBRE'] = df_visible['EDAD'].astype(int)
        df_visible.loc[df_visible['SEXO'] == 'F',
                       'MUJER'] = df_visible['EDAD'].astype(int)

        df_visible.loc[df_visible['REGIMEN'] ==
                       'Contributivo', 'CONTRIBUTIVO'] = 'X'
        df_visible.loc[df_visible['REGIMEN'] ==
                       'Subsidiado', 'SUBSIDIADO'] = 'X'

        for index, row in df_visible.iterrows():
            if (row['VICTIMA DE CONFLICTO'] == 'SI' or
                row['MIGRANTE'] == 'SI' or
                row['CARCELARIO'] == 'SI' or
                row['EN ESTADO DE GESTACION'] == 'SI' or
                row['OTRO'] == 'SI' or
                    row['DESPLAZADO'] == 'SI'):
                df_visible.at[index, 'CATEGORIA'] = 'X'

            else:
                df_visible.at[index, 'SINCATEGORIA'] = 'X'
        df_visible['MUJER'] = ''
        # Resto de la lógica de análisis

        output_file = os.path.join(app.config['UPLOAD_FOLDER'], file_end)
        df_visible.to_excel(output_file, index=False)
        flash('Análisis realizado con éxito.')
        return 1
    
    

    
